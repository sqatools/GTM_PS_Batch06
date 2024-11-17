import openpyxl
"""
# Open Pyxml data can be read from excel file 
"""

def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet_no = wb['Sheet1']
    print(sheet_no['A1'].value)


#read_excel("test_data.xlsx")
# write_excel


def write_excel(file_path, content):
    wb = openpyxl.load_workbook(file_path)
    sheet_no = wb['Sheet1']
    cell_name= sheet_no['B2']
    cell_name.value = content
    wb.save(file_path)


#write_excel("test_data.xlsx", "Hello Good monring")
#write_excel(300, 50)



def read_excel_with_name(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet_no = wb['Sheet1']
    for i in range(1, 5):
        print(sheet_no[f'A{i}'].value)

#read_excel_with_name("test_data.xlsx")


def read_excel_with_row_colns(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet_no = wb['Sheet2']
    max_row = sheet_no.max_row
    max_column = sheet_no.max_column
    for i in range(1, max_row+1):
        for j in range(1 , max_column+1):
            print(sheet_no.cell(row=i, column=j).value, end=" ")
        print()

read_excel_with_row_colns("test_data.xlsx")
